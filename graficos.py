from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

file_path = 'estoque.xlsx'
workbook = load_workbook(file_path)
sheet_estoque = workbook["Estoque"]

max_row = sheet_estoque.max_row
if sheet_estoque.cell(row=max_row, column=1).value == "Totais Gerais":
    max_row -= 1

if "Resumo" in workbook.sheetnames:
    del workbook["Resumo"]
sheet_resumo = workbook.create_sheet("Resumo")

sheet_resumo["A1"] = "Indicadores Gerais"
sheet_resumo["A3"] = "Total Geral de Valor"
sheet_resumo["B3"] = f'=SUM(Estoque!{get_column_letter(7)}2:{get_column_letter(7)}{max_row})'  # Coluna G

sheet_resumo["A4"] = "Total Geral de Lucro"
sheet_resumo["B4"] = f'=SUM(Estoque!{get_column_letter(6)}2:{get_column_letter(6)}{max_row})'  # Coluna F

sheet_resumo["A5"] = "Média de Lucratividade (%)"
sheet_resumo["B5"] = f'=AVERAGE(Estoque!{get_column_letter(3)}2:{get_column_letter(3)}{max_row})'  # Coluna C

sheet_resumo["A6"] = "Média de Quantidade em Estoque"
sheet_resumo["B6"] = f'=AVERAGE(Estoque!{get_column_letter(4)}2:{get_column_letter(4)}{max_row})'  # Coluna D

if "Gráficos" in workbook.sheetnames:
    del workbook["Gráficos"]
sheet_graficos = workbook.create_sheet("Gráficos")


produtos = Reference(sheet_estoque, min_col=1, min_row=2, max_row=max_row)
valores_totais = Reference(sheet_estoque, min_col=7, min_row=1, max_row=max_row)
lucratividade = Reference(sheet_estoque, min_col=3, min_row=1, max_row=max_row)
valor_fornecedor = Reference(sheet_estoque, min_col=2, min_row=1, max_row=max_row)
preco_venda = Reference(sheet_estoque, min_col=5, min_row=1, max_row=max_row)

bar_chart = BarChart()
bar_chart.title = "Valor Total por Produto"
bar_chart.add_data(valores_totais, titles_from_data=True)
bar_chart.set_categories(produtos)
bar_chart.y_axis.title = "Valor (R$)"
bar_chart.x_axis.title = "Produto"
bar_chart.width = 25
bar_chart.height = 10
sheet_graficos.add_chart(bar_chart, "A1")

line_chart = LineChart()
line_chart.title = "Lucratividade (%) por Produto"
line_chart.add_data(lucratividade, titles_from_data=True)
line_chart.set_categories(produtos)
line_chart.y_axis.title = "Lucratividade"
line_chart.x_axis.title = "Produto"
line_chart.width = 25
line_chart.height = 10
sheet_graficos.add_chart(line_chart, "A20")

pie_chart = PieChart()
pie_chart.title = "Top 5 Produtos por Valor"
pie_valores = Reference(sheet_estoque, min_col=7, min_row=2, max_row=min(max_row, 6))
pie_nomes = Reference(sheet_estoque, min_col=1, min_row=2, max_row=min(max_row, 6))
pie_chart.add_data(pie_valores, titles_from_data=False)
pie_chart.set_categories(pie_nomes)
pie_chart.width = 10
pie_chart.height = 10
sheet_graficos.add_chart(pie_chart, "A39")

empilhado_chart = BarChart()
empilhado_chart.title = "Preço de Venda x Valor Fornecedor"
empilhado_chart.type = "col"
empilhado_chart.grouping = "stacked"
empilhado_chart.add_data(valor_fornecedor, titles_from_data=True)
empilhado_chart.add_data(preco_venda, titles_from_data=True)
empilhado_chart.set_categories(produtos)
empilhado_chart.y_axis.title = "Valores (R$)"
empilhado_chart.x_axis.title = "Produto"
empilhado_chart.width = 30
empilhado_chart.height = 10
sheet_graficos.add_chart(empilhado_chart, "H1")

workbook.save(file_path)
